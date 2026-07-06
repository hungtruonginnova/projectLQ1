"""
Build the weekly Overdue Parts Report (Seawind).

Joins four production data exports into ONE row per (boat, material) showing
whether purchasing's ETAs will meet each boat's production need, then writes
the result into a copy of the report template so the output keeps the
template's layout / styles / autofilter.

The logic here was reverse-engineered from a real finished report (the WK27
"... PQ.xlsx" template, which the user confirmed is an example of correct
output). See references/instructions.md for the column-by-column mapping and
the evidence behind each filter.

Usage:
    python build_report.py \
        --opd "input1/OPD_Export ....xlsx" \
        --invmaster "input2/InvMaster_Export ....xlsx" \
        --workorders "input3/GAB_6271_OpenWorkOrders....xlsx" \
        --schedule "input4/Seawind production schedule....xlsx" \
        --template "template/OVERDUE PARTS REPORT WK 27 (29.06.2026) PQ.xlsx" \
        --out "output/OVERDUE PARTS REPORT WK 28 (06.07.2026) PQ.xlsx" \
        --report-date 2026-07-06 \
        --config config.yaml
"""
import argparse
import copy
import re
from datetime import date
from pathlib import Path

import pandas as pd
import yaml
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# --------------------------------------------------------------------------
# Helpers
# --------------------------------------------------------------------------

def load_config(path):
    with open(path, "r", encoding="utf-8") as f:
        cfg = yaml.safe_load(f)
    for key in ("drop_code_sorts", "critical_vendor_keywords",
                "critical_part_keywords", "local_vendor_keywords"):
        cfg[key] = [s.upper() for s in cfg.get(key, [])]
    return cfg


def norm_part(series):
    return series.astype(str).str.strip().str.upper()


def clean_text(series):
    """ERP exports embed literal _x000D_ carriage returns and trailing spaces
    in text cells - strip them so values are comparable and print cleanly."""
    return (series.astype(str)
            .str.replace("_x000D_", "", regex=False)
            .str.replace("\r", " ", regex=False)
            .str.replace("\n", " ", regex=False)
            .str.replace(r"\s+", " ", regex=True)
            .str.strip())


def contains_any(series, keywords):
    if not keywords:
        return pd.Series(False, index=series.index)
    pattern = "|".join(re.escape(k) for k in keywords)
    return series.astype(str).str.upper().str.contains(pattern, na=False, regex=True)


def join_unique(s):
    return " | ".join(dict.fromkeys(str(x).strip() for x in s.dropna() if str(x).strip()))


# --------------------------------------------------------------------------
# Readers
# --------------------------------------------------------------------------

def read_opd(path):
    cols = ["PURCHASE_ORDER", "SUPPLIER_NAME", "PART", "QTY_OPEN", "ETA_PO_LINE",
            "INTRANSIT", "ETA_PORT", "USER 2(DEPOSIT PAYMENT STATUS)",
            "USER 5 (BOAT NO./ URGENT/REVISED REASONS)", "ORIGINATOR", "STATUS",
            "PO_LINE_TEXT"]
    df = pd.read_excel(path, sheet_name=0, usecols=cols)
    df["PART_KEY"] = norm_part(df["PART"])
    df["ETA_PO_LINE"] = pd.to_datetime(df["ETA_PO_LINE"], errors="coerce")
    df["ETA_PORT"] = pd.to_datetime(df["ETA_PORT"], errors="coerce", format="mixed")
    df = df.rename(columns={
        "USER 2(DEPOSIT PAYMENT STATUS)": "USER2",
        "USER 5 (BOAT NO./ URGENT/REVISED REASONS)": "USER5",
    })
    for c in ["SUPPLIER_NAME", "ORIGINATOR", "USER2", "USER5", "INTRANSIT", "PO_LINE_TEXT"]:
        df[c] = clean_text(df[c]).replace("NAN", "")
    df = df[df["STATUS"].astype(str).str.upper().eq("OPEN")]
    return df


def read_invmaster(path):
    cols = ["PART", "CODE_SORT", "INACTIVE", "QTY_ONHAND", "QTY_ONORDER_PO",
            "VENDOR_NAME", "VENDOR_ADDRESS", "UM_INVENTORY"]
    df = pd.read_excel(path, sheet_name=0, usecols=cols)
    df["PART_KEY"] = norm_part(df["PART"])
    df = df.drop_duplicates("PART_KEY", keep="first")
    for c in ["VENDOR_NAME", "VENDOR_ADDRESS", "UM_INVENTORY"]:
        df[c] = clean_text(df[c]).replace("NAN", "")
    df["CODE_SORT"] = df["CODE_SORT"].astype(str).str.strip()
    return df


def read_workorders(path):
    cols = ["Customer PO", "Material", "Extra Desc", "Qty Rem.", "Job Seq Closed",
            "WO Mat Due Date", "WO Open Date", "Actual Material Generation Date"]
    df = pd.read_excel(path, sheet_name=0, usecols=cols)
    df["PART_KEY"] = norm_part(df["Material"])
    df["WO Mat Due Date"] = pd.to_datetime(df["WO Mat Due Date"], errors="coerce")
    df["WO Open Date"] = pd.to_datetime(df["WO Open Date"], errors="coerce")
    df["Actual Material Generation Date"] = pd.to_datetime(
        df["Actual Material Generation Date"], errors="coerce")
    df["Extra Desc"] = clean_text(df["Extra Desc"]).replace("NAN", "")

    # NOTE: do NOT drop Qty Rem. <= 0 lines here. Some Customer PO / material
    # combinations carry offsetting lines (e.g. +1 and -1) that must net out
    # together - dropping negative lines up front before the boat/material
    # rollup double-counts the positive lines instead of netting them. The
    # net-qty filter is applied after the BOM_QTY rollup in build().
    df = df[df["Job Seq Closed"].astype(str).str.upper().eq("N")]

    # WO ADJUST DATE = when material was actually generated for the WO, falling
    # back to when the WO was opened if generation hasn't happened (1900 = null).
    amg = df["Actual Material Generation Date"]
    df["WO_ADJUST"] = amg.where(amg.dt.year > 1900, df["WO Open Date"])

    cp = df["Customer PO"].astype(str).str.strip()
    parts = cp.str.split("-", n=1)
    df["BOAT"] = parts.str[0].str.strip()
    df["MODEL"] = parts.str[1].fillna("").str.strip()
    return df


def read_schedule(path):
    df = pd.read_excel(path, sheet_name=0, skiprows=2)
    df = df.rename(columns={df.columns[0]: "BOAT"})
    df["BOAT"] = df["BOAT"].astype(str).str.strip()
    df["BOAT_COMPLETION_DATE"] = pd.to_datetime(
        df["Planned boat completion date"], errors="coerce")
    keep = {"BOAT": "BOAT", "Model": "SCHED_MODEL",
            "Delivery option": "DELIVERY_OPTION", "BOAT_COMPLETION_DATE": "BOAT_COMPLETION_DATE"}
    df = df[[c for c in keep if c in df.columns]].rename(columns=keep)
    df = df.dropna(subset=["BOAT"]).drop_duplicates("BOAT", keep="first")
    for c in ["SCHED_MODEL", "DELIVERY_OPTION"]:
        if c in df.columns:
            df[c] = clean_text(df[c]).replace("NAN", "")
    return df


# --------------------------------------------------------------------------
# Build
# --------------------------------------------------------------------------

PEG_COLS = ["PURCHASE_ORDER", "ETA_PO_LINE", "INTRANSIT", "ETA_PORT",
            "QTY_OPEN", "ORIGINATOR", "USER5"]


def allocate_po(df, opd):
    """Allocate supply to boats FIFO by need date, mirroring how the buyer
    covers demand. For each material the supply is, in order:
      1. on-hand stock (`QTY_ONHAND`) - available immediately, no PO/ETA;
      2. open PO lines, earliest ETA first.
    Boats are sorted by need date (ETA REQUIRED = WO Mat Due Date) and walked
    top-down, each consuming its BOM_QTY from the running supply:
      - a boat whose demand begins while on-hand stock is still available is
        marked USING_STOCK (it needs no incoming PO);
      - the first boat past the on-hand quantity is pegged to the PO line that
        covers where its demand begins, and so on down the PO lines;
      - a boat whose demand falls beyond all supply gets no PO (a genuine,
        still-unordered shortage).

    This fixes the earlier bug where a boat that on-hand stock could cover was
    still pegged to a newly-incoming PO, because stock used to be checked only
    against total demand at the material level instead of allocated boat-by-boat
    in need-date order. Confirmed against the real WK27 report, this reproduces
    the per-boat assignment far better than giving every boat the 'soonest' PO."""
    buckets = {k: g for k, g in opd.sort_values("ETA_PO_LINE", na_position="last")
               .groupby("PART_KEY")}
    out = {c: [] for c in PEG_COLS}
    using_stock = []
    order = []
    for key, g in df.sort_values(["PART_KEY", "ETA_REQUIRED", "BOAT"]).groupby(
            "PART_KEY", sort=False):
        onhand = float(pd.to_numeric(g["QTY_ONHAND"], errors="coerce").fillna(0).iloc[0])
        b = buckets.get(key)
        supply = ([] if b is None else
                  list(zip(b["QTY_OPEN"].fillna(0).tolist(), b[PEG_COLS].to_dict("records"))))
        consumed = 0.0
        for ridx, row in g.iterrows():
            start = consumed          # where this boat's demand begins
            stock = start < onhand    # still inside the on-hand stock band
            chosen = None
            if not stock:
                pos = start - onhand   # position within the incoming-PO supply
                cum = 0.0
                for q, rec in supply:
                    if pos < cum + q or q == 0:
                        chosen = rec
                        break
                    cum += q
            order.append(ridx)
            using_stock.append(stock)
            for c in PEG_COLS:
                out[c].append(chosen[c] if chosen else None)
            consumed += row["BOM_QTY"] or 0
    res = pd.DataFrame(out, index=order).add_prefix("P_")
    res["USING_STOCK"] = pd.Series(using_stock, index=order)
    return df.join(res)


def build(cfg, opd, inv, wo, sched, report_date):
    report_date = pd.Timestamp(report_date)
    horizon = report_date + pd.DateOffset(months=int(cfg.get("horizon_months", 11)))
    earliest = report_date - pd.DateOffset(days=int(cfg.get("done_buffer_days", 45)))
    drop_log = []

    def drop(mask, reason, df):
        n = int(mask.sum())
        if n:
            drop_log.append((reason, n))
        return df[~mask]

    # --- clean OPD supply before any PO matching --------------------------
    # Cancelled / no-real-qty PO lines must never be pegged to a boat or shown
    # in the GSS backup block:
    #  - QTY_OPEN <= sentinel (0.0001): closed/cancelled/dropship/NCR sentinel
    #  - note text (USER5 / PO_LINE_TEXT) says the buyer cancelled the line or
    #    raised an NCR (\bNCR\b word-boundary so "increase" isn't caught).
    sentinel = float(cfg.get("qty_open_sentinel", 0.0001))
    note = opd["USER5"].astype(str) + " " + opd["PO_LINE_TEXT"].astype(str)
    cancelled = note.str.contains(r"cancel|\bNCR\b", case=False, regex=True, na=False)
    n_opd0 = len(opd)
    opd = opd[(opd["QTY_OPEN"].fillna(0) > sentinel) & ~cancelled]
    print(f"  OPD lines removed (qty<={sentinel} or cancel/NCR note): "
          f"{n_opd0 - len(opd)}")

    # --- collapse work orders to one row per (boat, material) --------------
    # BOM QTY is the summed outstanding qty; the dates come from the single
    # most-urgent line (earliest WO Mat Due Date) so ETA REQUIRED / WO ADJUST
    # stay a consistent pair from one work order.
    wo_sorted = wo.sort_values("WO Mat Due Date", na_position="last")
    rep = wo_sorted.drop_duplicates(["BOAT", "PART_KEY"], keep="first")[
        ["BOAT", "PART_KEY", "Material", "Extra Desc", "MODEL",
         "WO Mat Due Date", "WO_ADJUST"]].rename(columns={
             "Extra Desc": "Extra_Desc", "WO Mat Due Date": "ETA_REQUIRED"})
    bom = wo.groupby(["BOAT", "PART_KEY"], as_index=False)["Qty Rem."].sum().rename(
        columns={"Qty Rem.": "BOM_QTY"})
    grp = rep.merge(bom, on=["BOAT", "PART_KEY"], how="left")

    # net-qty filter: offsetting lines (e.g. +1 and -1 for the same boat/
    # material) must cancel out BEFORE the row survives, not be counted twice.
    grp = drop(grp["BOM_QTY"].fillna(0) <= 0,
               "Qty Rem. nets to zero/negative across Customer PO lines (fully issued/cancelled)",
               grp)

    # Total demand per part = sum of BOM_QTY across ALL open boats needing
    # that material, per the source instructions - computed here, before the
    # schedule/horizon filters below drop rows, so it isn't undercounted by
    # boats that fall outside this report's window but still hold real stock.
    total_demand = grp.groupby("PART_KEY")["BOM_QTY"].sum().rename("TOTAL_DEMAND")

    # --- boat filter: must be a scheduled boat within the planning window --
    df = grp.merge(sched, on="BOAT", how="left")
    df = drop(df["BOAT_COMPLETION_DATE"].isna(),
              "Boat not in production schedule (or no completion date)", df)
    df = drop(df["BOAT_COMPLETION_DATE"] > horizon,
              f"Boat completes beyond planning horizon (> {cfg.get('horizon_months',11)} months)", df)
    df = drop(df["BOAT_COMPLETION_DATE"] < earliest,
              "Boat already completed / shipped (before window)", df)

    # --- enrich from InvMaster --------------------------------------------
    inv_cols = inv.set_index("PART_KEY")[
        ["CODE_SORT", "INACTIVE", "QTY_ONHAND", "QTY_ONORDER_PO",
         "VENDOR_NAME", "VENDOR_ADDRESS", "UM_INVENTORY"]]
    df = df.merge(inv_cols, on="PART_KEY", how="left")

    df = drop(df["INACTIVE"].fillna(False).astype(bool), "Inactive part (InvMaster)", df)

    # Total demand column (H): total open demand for the material across all
    # boats. Kept for reference; on-hand stock is now allocated boat-by-boat in
    # allocate_po() (below) instead of compared against this total.
    df = df.merge(total_demand, on="PART_KEY", how="left")

    # --- consumable / local-vendor drops -----------------------------------
    # Consumables / auxiliary materials (CODE_SORT in the drop list) are removed
    # outright - no critical exception - per the buyer's instruction that these
    # groups never belong in this report.
    consumable = df["CODE_SORT"].astype(str).str.upper().isin(cfg["drop_code_sorts"])
    df = drop(consumable, "Consumable / auxiliary material code-sort", df)

    # Local (Vietnam) vendor parts still keep the critical-keyword exception.
    critical = (contains_any(df["VENDOR_NAME"], cfg["critical_vendor_keywords"])
                | contains_any(df["Material"], cfg["critical_part_keywords"])
                | contains_any(df["Extra_Desc"], cfg["critical_part_keywords"]))
    local = contains_any(df["VENDOR_ADDRESS"], cfg["local_vendor_keywords"])
    df = drop(local & ~critical, "Local (Vietnam) vendor part (non-critical)", df)

    # --- enrich from OPD: allocate on-hand stock + PO lines to boats -------
    # On-hand stock is consumed first by the earliest-needed boats; remaining
    # boats are pegged to incoming PO lines in ETA order. Boats fully covered by
    # on-hand stock are dropped ("using stock").
    df = allocate_po(df, opd)
    df = drop(df["USING_STOCK"].fillna(False).astype(bool),
              "On-hand stock covers earlier boats (using stock, by need date)", df)

    if len(opd):
        gss = opd.groupby("PART_KEY").agg(
            PO_GSS=("PURCHASE_ORDER", join_unique),
            VENDOR_GSS=("SUPPLIER_NAME", join_unique),
            QTY_GSS=("QTY_OPEN", join_unique),
            ETA_GSS=("ETA_PO_LINE", lambda s: join_unique(
                s.dropna().dt.strftime("%d/%m/%Y"))),
            USER2_GSS=("USER2", join_unique),
            USER5_GSS=("USER5", join_unique),
            PO_LINE_TEXT_GSS=("PO_LINE_TEXT", join_unique),
        )
        df = df.merge(gss, on="PART_KEY", how="left")
    else:
        for c in ["PO_GSS", "VENDOR_GSS", "QTY_GSS", "ETA_GSS", "USER2_GSS",
                  "USER5_GSS", "PO_LINE_TEXT_GSS"]:
            df[c] = pd.NA

    # --- GAP = ETA REQUIRED - ETA (days); blank when no ETA ----------------
    df["GAP"] = (df["ETA_REQUIRED"] - df["P_ETA_PO_LINE"]).dt.days

    # --- review flags (kept, not dropped) ---------------------------------
    df["REVIEW"] = False
    df["NOTE"] = ""

    def flag(mask, reason):
        mask = mask.reindex(df.index, fill_value=False)
        df.loc[mask, "REVIEW"] = True
        cur = df.loc[mask, "NOTE"]
        df.loc[mask, "NOTE"] = cur.where(cur == "", cur + "; ") + reason

    # recompute on the post-drop / post-merge frame so masks align to df.index
    flag(df["P_PURCHASE_ORDER"].isna(), "No open PO found - verify sourcing")
    flag(df["P_ETA_PORT"].notna() & (df["P_ETA_PORT"] <= report_date),
         "Already at port/airport - confirm & remove if cleared")
    flag(df["GAP"].notna() & (df["GAP"] <= int(cfg.get("gap_days_not_urgent", 7))),
         f"Gap <= {cfg.get('gap_days_not_urgent',7)} days - candidate to remove")

    return df, drop_log


# --------------------------------------------------------------------------
# Output
# --------------------------------------------------------------------------

def build_boat_label(row):
    if pd.notna(row["BOAT_COMPLETION_DATE"]):
        dstr = row["BOAT_COMPLETION_DATE"].strftime("%d-%b-%y")
    else:
        dstr = "TBC"

    def clean(v):
        s = "" if v is None else str(v).strip()
        return "" if s.lower() in ("", "nan", "nat", "none") else s

    model = clean(row.get("MODEL")) or clean(row.get("SCHED_MODEL"))
    # Delivery option comes from input4; blank (e.g. trimarans) -> TBC, never "nan".
    opt = clean(row.get("DELIVERY_OPTION")) or "TBC"
    return f"{row['BOAT']}-{model}- planned_completion-{dstr}-{opt}, "


def to_report_frame(df):
    out = pd.DataFrame(index=df.index)
    out["BOAT"] = df["BOAT"]
    out["BOAT_LABEL"] = df.apply(build_boat_label, axis=1)
    out["Material"] = df["Material"]
    out["Extra Desc"] = df["Extra_Desc"]
    out["Inv UM"] = df["UM_INVENTORY"]
    out["Code Sort"] = df["CODE_SORT"]
    out["BOM QTY"] = df["BOM_QTY"]
    out["Total"] = df["TOTAL_DEMAND"]
    out["On Hand"] = df["QTY_ONHAND"]
    out["Qty OnOrder"] = df["QTY_ONORDER_PO"]
    out["Default Vendor"] = df["VENDOR_NAME"]
    out["Buyer"] = df["P_ORIGINATOR"]
    out["ETA"] = df["P_ETA_PO_LINE"]
    out["KEY"] = df["Material"].astype(str) + df["BOAT"].astype(str)
    out["ETA REQUIRED"] = df["ETA_REQUIRED"]
    out["BOAT COMPLETION DATE"] = df["BOAT_COMPLETION_DATE"]
    out["GAP"] = df["GAP"]
    out["WO ADJUST DATE"] = df["WO_ADJUST"]
    out["PO"] = df["P_PURCHASE_ORDER"]
    out["IN TRANSIT"] = df["P_INTRANSIT"]
    out["ETA PORT"] = df["P_ETA_PORT"]
    out["Quantity Coming"] = df["P_QTY_OPEN"]
    out["REMARK"] = df["P_USER5"]
    out["REMARK2"] = ""
    out["NOTE"] = df["NOTE"]
    out["PO Number GSS"] = df["PO_GSS"]
    out["Vendor (OPD)"] = df["VENDOR_GSS"]
    out["Qty on PO GSS"] = df["QTY_GSS"]
    out["ETA GSS"] = df["ETA_GSS"]
    out["USERFIELD 2 GSS"] = df["USER2_GSS"]
    out["USERFIELD 5 GSS"] = df["USER5_GSS"]
    out["PO_LINE_TEXT"] = df["PO_LINE_TEXT_GSS"]
    out["_REVIEW"] = df["REVIEW"].values
    out = out.sort_values(["GAP", "BOAT"], na_position="last")
    return out


DATE_COLS = {13, 15, 16, 18, 21}  # 1-based: ETA, ETA REQUIRED, BOAT COMPLETION, WO ADJUST, ETA PORT
DATE_FORMAT = "dd/mm/yyyy"
REVIEW_FILL = "FFFF00"


def write_output(template_path, out_path, report_df):
    wb = load_workbook(template_path)
    ws = wb["original (2)"]

    # the user only wants the visible report sheet - drop the hidden rollup
    if "original" in wb.sheetnames:
        del wb["original"]

    header_row, first_data_row = 3, 4
    max_row = ws.max_row
    style_cells = {c.column_letter: copy.copy(c._style) for c in ws[first_data_row]}
    if max_row >= first_data_row:
        ws.delete_rows(first_data_row, max_row - first_data_row + 1)

    review_fill = PatternFill(start_color=REVIEW_FILL, end_color=REVIEW_FILL, fill_type="solid")

    def cell_value(v):
        if isinstance(v, pd.Timestamp):
            return v.to_pydatetime() if pd.notna(v) else None
        if pd.isna(v):
            return None
        return v

    review_flags = report_df["_REVIEW"].tolist()
    data = report_df.drop(columns=["_REVIEW"])
    for i, (_, row) in enumerate(data.iterrows()):
        r = first_data_row + i
        is_review = review_flags[i]
        for j, value in enumerate(row.tolist(), start=1):
            cell = ws.cell(row=r, column=j, value=cell_value(value))
            letter = cell.column_letter
            if letter in style_cells:
                cell._style = copy.copy(style_cells[letter])
            if j in DATE_COLS and isinstance(cell.value, __import__("datetime").datetime):
                cell.number_format = DATE_FORMAT
            if is_review:
                cell.fill = review_fill

    last_row = max(first_data_row + len(data) - 1, header_row)
    if ws.auto_filter.ref:
        a, b = ws.auto_filter.ref.split(":")
        col_a = "".join(ch for ch in a if ch.isalpha())
        col_b = "".join(ch for ch in b if ch.isalpha())
        ws.auto_filter.ref = f"{col_a}{header_row}:{col_b}{last_row}"

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


# --------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------

def generate_report(opd_path, invmaster_path, workorders_path, schedule_path,
                    template_path, out_path, report_date=None, config_path=None):
    """Programmatic entry point for web UI / API wrappers."""
    cfg = load_config(config_path or Path(__file__).parent / "config.yaml")
    report_date = report_date or date.today().isoformat()

    opd = read_opd(opd_path)
    inv = read_invmaster(invmaster_path)
    wo = read_workorders(workorders_path)
    sched = read_schedule(schedule_path)

    df, drop_log = build(cfg, opd, inv, wo, sched, report_date)
    report = to_report_frame(df)
    write_output(template_path, out_path, report)

    return {
        "input_counts": {
            "opd": len(opd),
            "invmaster": len(inv),
            "workorders": len(wo),
            "schedule": len(sched),
        },
        "drop_log": drop_log,
        "row_count": len(report),
        "review_count": int(report["_REVIEW"].sum()),
        "report_date": report_date,
        "out_path": str(out_path),
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--opd", required=True)
    ap.add_argument("--invmaster", required=True)
    ap.add_argument("--workorders", required=True)
    ap.add_argument("--schedule", required=True)
    ap.add_argument("--template", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--report-date", default=None,
                    help="YYYY-MM-DD; drives the boat window & 'already at port'. Defaults to today.")
    ap.add_argument("--config", default=str(Path(__file__).parent / "config.yaml"))
    args = ap.parse_args()

    cfg = load_config(args.config)
    report_date = args.report_date or date.today().isoformat()

    print("Reading inputs...")
    opd = read_opd(args.opd)
    inv = read_invmaster(args.invmaster)
    wo = read_workorders(args.workorders)
    sched = read_schedule(args.schedule)
    print(f"  OPD open PO lines : {len(opd)}")
    print(f"  InvMaster parts   : {len(inv)}")
    print(f"  6271 open WO lines : {len(wo)}")
    print(f"  Schedule boats    : {len(sched)}")

    df, drop_log = build(cfg, opd, inv, wo, sched, report_date)
    report = to_report_frame(df)

    print("\nRows removed by rule:")
    total = 0
    for reason, n in drop_log:
        print(f"  {n:>6}  {reason}")
        total += n
    print(f"  {total:>6}  TOTAL removed")

    n_rev = int(report["_REVIEW"].sum())
    print(f"\nFinal report rows: {len(report)}  (flagged for review: {n_rev})")
    print(f"Report date used : {report_date}")

    write_output(args.template, args.out, report)
    print(f"\nWrote {args.out}")


if __name__ == "__main__":
    main()
